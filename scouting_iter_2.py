#Scouting Project
#Sagar Patel - 211Z
#Began on Friday June 19, 2020
#Description: Scouting System Between Computers (No Wifi)
#--------- MATCHES ------------#
#Blue 1 For Match 1: 2381Z
#Blue 1 For Match 2: 839Z
#Blue 1 For Match 3: 2381Z
#------------------------------#
from openpyxl import load_workbook
class Team:
    def __init__(self,teamname=None,auton=None,ownmiddle=None,contribution=None,finesse=None,drivetype=None,drivespeed=None,intakespeed=None,
                 ballctrl=None,scoringspeed=None,defencerating=None,winlose=None,filename = None, wb=None,ws=None):
        self.teamname = teamname
        self.auton = auton
        self.ownmiddle = ownmiddle
        self.contribution = contribution
        self.finesse = finesse
        self.drivetype = drivetype
        self.drivespeed = drivespeed
        self.intakespeed = intakespeed
        self.ballctrl = ballctrl
        self.scoringspeed = scoringspeed
        self.defencerating = defencerating
        self.winlose = winlose
        self.filename = filename
        self.wb = wb
        self.ws = ws
    def scout_input(self):
        self.auton.append(input("Enter Auton Points: "))
        self.ownmiddle.append(input("Enter Middle Goal Owned: "))
        self.contribution.append(input("Enter Team Contribution: "))
        self.finesse.append(input("Enter Driver Finesse: "))
        self.drivetype.append(input("Enter Drive Type: "))
        self.drivespeed.append(input("Enter Drive Speed: "))
        self.intakespeed.append(input("Enter Intake Speed: "))
        self.ballctrl.append(input("Enter Ball Control: "))
        self.scoringspeed.append(input("Enter Scoring Speed: "))
        self.defencerating.append(input("Enter Defence Rating: "))
        self.winlose.append(input("Enter Win or Lose: "))
    def excelupload(self):
        self.filename = str(self.teamname) + ".xlsx"
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.worksheets[0]
        self.ws_tables = []
        rowvalue = 2
        for item in   self.auton :
            self.ws["A" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.ownmiddle :
            self.ws["B" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.contribution :
            self.ws["C" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.finesse :
            self.ws["D" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.drivetype :
            self.ws["E" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.drivespeed :
            self.ws["F" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.intakespeed :
            self.ws["G" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.ballctrl :
            self.ws["H" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.scoringspeed :
            self.ws["I" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.defencerating :
            self.ws["J" + str(rowvalue)] = str(item)
            rowvalue += 1
        rowvalue = 2
        for item in self.winlose :
            self.ws["K" + str(rowvalue)] = str(item)
            rowvalue += 1
        self.wb.save(self.filename)
#---------- TEAMS -------------#
team839Z = Team ("839Z",[],[],[],[],[],[],[],[],[],[],[],None,None,None)
team2381Z = Team ("2381Z",[],[],[],[],[],[],[],[],[],[],[],None,None,None)
#--------- MATCH 1 ------------#
print ("Match 1 - 2381Z")
team2381Z.scout_input()
team2381Z.excelupload()
#--------- MATCH 2 ------------#
print ("Match 2 - 839Z")
team839Z.scout_input()
team839Z.excelupload()
#--------- MATCH 3 ------------#
print ("Match 3 - 2381Z")
team2381Z.scout_input()
team2381Z.excelupload()


