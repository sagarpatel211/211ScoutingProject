#Scouting Project
#Sagar Patel - 211Z
#Began on Friday June 19, 2020
#Description: Scouting System Between Computers (No Wifi)
#--------- MATCHES ------------#
#Blue 1 For Match 1: 2381Z
#Blue 1 For Match 2: 839Z
#Blue 1 For Match 3: 2381Z
#------------------------------#
from openpyxl import load_workbook
from tkinter import *
import tkinter.messagebox as tmsg
window = Tk()
window.title('SCOUTING')
window.geometry("400x600+0+0")
window.resizable(False, False)
class Team:
    def __init__(self,teamname,auton,ownmiddle,contribution,finesse,drivetype,drivespeed,intakespeed,
                 ballctrl,scoringspeed,defencerating,winlose,filename,wb,ws):
        self.teamname = teamname
        self.auton = auton
        self.ownmiddle = ownmiddle
        self.contribution = contribution
        self.finesse = finesse
        self.drivetype = drivetype
        self.drivespeed = drivespeed
        self.intakespeed = intakespeed
        self.ballctrl = ballctrl
        self.scoringspeed = scoringspeed
        self.defencerating = defencerating
        self.winlose = winlose
        self.filename = filename
        self.wb = wb
        self.ws = ws
    def display (self) :
        self.maintitle=Label( text="TEAM NAME - " + str(self.teamname), fg='black', font=("Helvetica", 20))
        self.maintitle.place(x=5, y=5)
        self.AutonPointsText=Label( text="How many auton points?", fg='black', font=("Helvetica", 10))
        self.AutonPointsText.place(x=10, y=50)
        self.AutonPointsInput=Entry()
        self.AutonPointsInput.place(x=230, y=50)
        self.OwnedMiddleText=Label( text="Do They Own Middle Tower?(Y/N)", fg='black', font=("Helvetica", 10))
        self.OwnedMiddle=Entry()
        self.OwnedMiddleText.place(x=10, y=95)
        self.OwnedMiddle.place(x=230,y=95)
        self.ContributionText=Label( text="How much did they contribute?(1-10)", fg='black', font=("Helvetica", 10))
        self.ContributionInput=Entry()
        self.ContributionText.place(x=10,y=140)
        self.ContributionInput.place(x=230, y=140)
        self.FinesseText=Label( text="How good is the driver?(1-10)", fg='black', font=("Helvetica", 10))
        self.FinesseInput=Entry()
        self.FinesseText.place(x=10,y=185)
        self.FinesseInput.place(x=230, y=185)
        self.DriveTypeText=Label(text="What's the drive type?", fg='black', font=("Helvetica", 10))
        self.DriveTypeText.place(x=10, y=230)
        self.DriveTypeInput=Entry()
        self.DriveTypeInput.place(x=230, y=230)
        self.DriveSpeedText=Label( text="How fast is the drivebase?(1-10)", fg='black', font=("Helvetica", 10))
        self.DriveSpeedInput=Entry()
        self.DriveSpeedText.place(x=10,y=275)
        self.DriveSpeedInput.place(x=230, y=275)
        self.IntakeSpeedText=Label( text="How fast are the intakes?(1-10)", fg='black', font=("Helvetica", 10))
        self.IntakeSpeedInput=Entry()
        self.IntakeSpeedText.place(x=10,y=320)
        self.IntakeSpeedInput.place(x=230, y=320)
        self.BallCtrlText=Label( text="How well do they control balls?(1-10)", fg='black', font=("Helvetica", 10))
        self.BallCtrlInput=Entry()
        self.BallCtrlText.place(x=10,y=365)
        self.BallCtrlInput.place(x=230, y=365)
        self.ScoringSpeedText=Label( text="How fast do they score?(1-10)", fg='black', font=("Helvetica", 10))
        self.ScoringSpeedInput=Entry()
        self.ScoringSpeedText.place(x=10,y=410)
        self.ScoringSpeedInput.place(x=230, y=410)
        self.DefenceText=Label( text="How well do they defend?(1-10)", fg='black', font=("Helvetica", 10))
        self.DefenceInput=Entry()
        self.DefenceText.place(x=10,y=455)
        self.DefenceInput.place(x=230, y=455)
        self.WinLoseText=Label( text="Did They Win or Lose?(W/L)", fg='black', font=("Helvetica", 10))
        self.WinLose = Entry()
        self.WinLoseText.place(x=10, y=500)
        self.WinLose.place(x=230,y=500)
        self.SubmitButton = Button(text='SUBMIT!', command=self.submitscoutinginfo)
        self.SubmitButton.place(x=250, y=550)
    def submitscoutinginfo(self):
        self.auton.append(self.AutonPointsInput.get())
        self.ownmiddle.append(self.OwnedMiddle.get())
        self.contribution.append(self.ContributionInput.get())
        self.finesse.append(self.FinesseInput.get())
        self.drivetype.append(self.DriveTypeInput.get())
        self.drivespeed.append(self.DriveSpeedInput.get())
        self.intakespeed.append(self.IntakeSpeedInput.get())
        self.ballctrl.append(self.BallCtrlInput.get())
        self.scoringspeed.append(self.ScoringSpeedInput.get())
        self.defencerating.append(self.DefenceInput.get())
        self.winlose.append(self.WinLose.get())
        self.filename = str(self.teamname) + ".xlsx"
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.worksheets[0]
        self.ws_tables = []
        self.rowvalue = 1
        for item in   self.auton :
            self.rowvalue += 1
            self.ws["A" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.ownmiddle :
            self.rowvalue += 1
            self.ws["B" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.contribution :
            self.rowvalue += 1
            self.ws["C" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.finesse :
            self.rowvalue += 1
            self.ws["D" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.drivetype :
            self.rowvalue += 1
            self.ws["E" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.drivespeed :
            self.rowvalue += 1
            self.ws["F" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.intakespeed :
            self.rowvalue += 1
            self.ws["G" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.ballctrl :
            self.rowvalue += 1
            self.ws["H" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.scoringspeed :
            self.rowvalue += 1
            self.ws["I" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.defencerating :
            self.rowvalue += 1
            self.ws["J" + str(self.rowvalue)] = str(item)
        self.rowvalue = 1
        for item in self.winlose :
            self.rowvalue += 1
            self.ws["K" + str(self.rowvalue)] = str(item)
        self.wb.save(self.filename)
#---------- TEAMS -------------#
team2381Z = Team ("2381Z",[],[],[],[],[],[],[],[],[],[],[],None,None,None)
team839Z = Team ("839Z",[],[],[],[],[],[],[],[],[],[],[],None,None,None)

#-- TKINTER DISPLAY FUNCTION --#
def display (teamname):
    window = Tk()
    window.title('SCOUTING')
    window.geometry("400x600+0+0")
    window.resizable(False, False)
    teamname.display()
    teamname.submitscoutinginfo()
    window.mainloop()
#--------- MATCHES ------------#
#1 - 2381Z
team2381Z.display()
team2381Z.submitscoutinginfo()
window.mainloop()

#2 - 839Z
display(team839Z)

#3 - 2381Z
display(team2381Z)
