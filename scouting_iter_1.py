# Scouting Project
#Sagar Patel - 211Z
#Began on Friday June 19, 2020
#Description: Scouting System Between Computers (No Wifi)
#############MATCHES###############
#Blue 1 For Match 1: 2381Z
#Blue 1 For Match 2: 839Z
#Blue 1 For Match 3: 2381Z
###################################
from openpyxl import load_workbook
import tkinter
AutoPoints2381Z = []
OwnedMiddleGoal2381Z = [] #Yes/No
ContributionRating2381Z = []
DriverFinesse2381Z = []
DriveType2381Z = []
DriveSpeed2381Z = []
IntakeSpeed2381Z = []
BallControl2381Z= []
ScoringSpeed2381Z = []
DefenceRating2381Z = []
WinOrLose2381Z = []

AutoPoints839Z = []
OwnedMiddleGoal839Z = [] #Yes/No
ContributionRating839Z = []
DriverFinesse839Z = []
DriveType839Z = []
DriveSpeed839Z = []
IntakeSpeed839Z = []
BallControl839Z= []
ScoringSpeed839Z = []
DefenceRating839Z = []
WinOrLose839Z = []
############## MATCH 1 ##############
print ("Match 1 - 2381Z")
InputAutoPoints2381Z = input("Enter Auton Points: ")
InputOwnedMiddleGoal2381Z = input("Enter Middle Goal Owned: ")
InputContributionRating2381Z = input("Enter Team Contribution: ")
InputDriverFinesse2381Z = input("Enter Driver Finesse: ")
InputDriveType2381Z = input("Enter Drive Type: ")
InputDriveSpeed2381Z = input("Enter Drive Speed: ")
InputIntakeSpeed2381Z = input("Enter Intake Speed: ")
InputBallControl2381Z= input("Enter Ball Control: ")
InputScoringSpeed2381Z = input("Enter Scoring Speed: ")
InputDefenceRating2381Z = input("Enter Defence Rating: ")
InputWinOrLose2381Z = input("Enter Win or Lose: ")

AutoPoints2381Z.append(InputAutoPoints2381Z)
OwnedMiddleGoal2381Z.append(InputOwnedMiddleGoal2381Z)
ContributionRating2381Z.append(InputContributionRating2381Z)
DriverFinesse2381Z.append(InputDriverFinesse2381Z)
DriveType2381Z.append(InputDriveType2381Z)
DriveSpeed2381Z.append(InputDriveSpeed2381Z)
IntakeSpeed2381Z.append(InputIntakeSpeed2381Z)
BallControl2381Z.append(InputBallControl2381Z)
ScoringSpeed2381Z.append(InputScoringSpeed2381Z)
DefenceRating2381Z.append(InputDefenceRating2381Z)
WinOrLose2381Z.append(InputWinOrLose2381Z)

filename2381Z = "2381Z.xlsx"
wb2381Z = load_workbook(filename2381Z)
ws2381Z = wb2381Z.worksheets[0]
ws2381Z_tables = []
rowvalue = 2
for item in AutoPoints2381Z :
    ws2381Z["A" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in OwnedMiddleGoal2381Z :
    ws2381Z["B" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in ContributionRating2381Z :
    ws2381Z["C" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriverFinesse2381Z :
    ws2381Z["D" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriveType2381Z :
    ws2381Z["E" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriveSpeed2381Z :
    ws2381Z["F" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in IntakeSpeed2381Z :
    ws2381Z["G" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in BallControl2381Z :
    ws2381Z["H" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in ScoringSpeed2381Z :
    ws2381Z["I" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DefenceRating2381Z :
    ws2381Z["J" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in WinOrLose2381Z :
    ws2381Z["K" + str(rowvalue)] = str(item)
    rowvalue += 1
wb2381Z.save(filename2381Z)
############## MATCH 2 ##############
print ("Match 2 - 839Z")
InputAutoPoints839Z = input("Enter Auton Points: ")
InputOwnedMiddleGoal839Z = input("Enter Middle Goal Owned: ")
InputContributionRating839Z = input("Enter Team Contribution: ")
InputDriverFinesse839Z = input("Enter Driver Finesse: ")
InputDriveType839Z = input("Enter Drive Type: ")
InputDriveSpeed839Z = input("Enter Drive Speed: ")
InputIntakeSpeed839Z = input("Enter Intake Speed: ")
InputBallControl839Z= input("Enter Ball Control: ")
InputScoringSpeed839Z = input("Enter Scoring Speed: ")
InputDefenceRating839Z = input("Enter Defence Rating: ")
InputWinOrLose839Z = input("Enter Win or Lose: ")

AutoPoints839Z.append(InputAutoPoints839Z)
OwnedMiddleGoal839Z.append(InputOwnedMiddleGoal839Z)
ContributionRating839Z.append(InputContributionRating839Z)
DriverFinesse839Z.append(InputDriverFinesse839Z)
DriveType839Z.append(InputDriveType839Z)
DriveSpeed839Z.append(InputDriveSpeed839Z)
IntakeSpeed839Z.append(InputIntakeSpeed839Z)
BallControl839Z.append(InputBallControl839Z)
ScoringSpeed839Z.append(InputScoringSpeed839Z)
DefenceRating839Z.append(InputDefenceRating839Z)
WinOrLose839Z.append(InputWinOrLose839Z)

filename839Z = "839Z.xlsx"
wb839Z = load_workbook(filename839Z)
ws839Z = wb839Z.worksheets[0]
ws839Z_tables = []
rowvalue = 2
for item in AutoPoints839Z :
    ws839Z["A" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in OwnedMiddleGoal839Z :
    ws839Z["B" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in ContributionRating839Z :
    ws839Z["C" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriverFinesse839Z :
    ws839Z["D" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriveType839Z :
    ws839Z["E" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriveSpeed839Z :
    ws839Z["F" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in IntakeSpeed839Z :
    ws839Z["G" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in BallControl839Z :
    ws839Z["H" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in ScoringSpeed839Z :
    ws839Z["I" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DefenceRating839Z :
    ws839Z["J" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in WinOrLose839Z :
    ws839Z["K" + str(rowvalue)] = str(item)
    rowvalue += 1
wb839Z.save(filename839Z)
############## MATCH 3 ##############
print ("Match 3 - 2381Z")
InputAutoPoints2381Z = input("Enter Auton Points: ")
InputOwnedMiddleGoal2381Z = input("Enter Middle Goal Owned: ")
InputContributionRating2381Z = input("Enter Team Contribution: ")
InputDriverFinesse2381Z = input("Enter Driver Finesse: ")
InputDriveType2381Z = input("Enter Drive Type: ")
InputDriveSpeed2381Z = input("Enter Drive Speed: ")
InputIntakeSpeed2381Z = input("Enter Intake Speed: ")
InputBallControl2381Z= input("Enter Ball Control: ")
InputScoringSpeed2381Z = input("Enter Scoring Speed: ")
InputDefenceRating2381Z = input("Enter Defence Rating: ")
InputWinOrLose2381Z = input("Enter Win or Lose: ")

AutoPoints2381Z.append(InputAutoPoints2381Z)
OwnedMiddleGoal2381Z.append(InputOwnedMiddleGoal2381Z)
ContributionRating2381Z.append(InputContributionRating2381Z)
DriverFinesse2381Z.append(InputDriverFinesse2381Z)
DriveType2381Z.append(InputDriveType2381Z)
DriveSpeed2381Z.append(InputDriveSpeed2381Z)
IntakeSpeed2381Z.append(InputIntakeSpeed2381Z)
BallControl2381Z.append(InputBallControl2381Z)
ScoringSpeed2381Z.append(InputScoringSpeed2381Z)
DefenceRating2381Z.append(InputDefenceRating2381Z)
WinOrLose2381Z.append(InputWinOrLose2381Z)

filename2381Z = "2381Z.xlsx"
wb2381Z = load_workbook(filename2381Z)
ws2381Z = wb2381Z.worksheets[0]
ws2381Z_tables = []
rowvalue = 2
for item in AutoPoints2381Z :
    ws2381Z["A" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in OwnedMiddleGoal2381Z :
    ws2381Z["B" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in ContributionRating2381Z :
    ws2381Z["C" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriverFinesse2381Z :
    ws2381Z["D" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriveType2381Z :
    ws2381Z["E" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DriveSpeed2381Z :
    ws2381Z["F" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in IntakeSpeed2381Z :
    ws2381Z["G" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in BallControl2381Z :
    ws2381Z["H" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in ScoringSpeed2381Z :
    ws2381Z["I" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in DefenceRating2381Z :
    ws2381Z["J" + str(rowvalue)] = str(item)
    rowvalue += 1
rowvalue = 2
for item in WinOrLose2381Z :
    ws2381Z["K" + str(rowvalue)] = str(item)
    rowvalue += 1
wb2381Z.save(filename2381Z)
##########################################################
